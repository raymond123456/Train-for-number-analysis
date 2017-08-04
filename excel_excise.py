import openpyxl,pprint
wb = openpyxl.load_workbook("C:\Users\RaymondWang\Desktop/censuspopdata.xlsx") #open excel,get a excel class object
#print wb.get_sheet_names() #objet's function
#sheet = wb.get_sheet_by_name('Sheet3')#(sheet type calss)
#a = sheet['B1'] #a type Cell
#print "Row" + str(a.row) + ', Column' + str(a.column)+' is ' + str(a.value)#a The cell attribute
#print a.value

#print sheet.cell(row=1,column=2).value#get B1 = sheet(B1)
#print sheet.get_highest_row()
#for i in range(1,4):
#   sheet_all = wb.get_sheet_by_name('Sheet'+str(i))
#   for j in range(1,3):
#        for k in range(1,10):
#           print sheet_all.cell(row=k,column=j).value
#        print "\n"
sheet = wb.get_sheet_by_name('Population by Census Tract')
countydate = {}
for row in range(2, sheet.max_row+1):
    state = sheet['B'+ str(row)].value
    county = sheet['C'+ str(row)].value
    pop = sheet['D'+str(row)].value
#set key default
    countydate.setdefault(state,{})
# make sure default
    countydate[state].setdefault(county, {'tracts': 0, 'pop': 0})
    countydate[state][county]['tracts'] += 1
    countydate[state][county]['pop'] += int(pop)
resultfile = open('D:\census2010.txt', 'w')










