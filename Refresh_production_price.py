import openpyxl
wb = openpyxl.load_workbook('C:\Users\RaymondWang\Desktop/produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')
"""
for row in range(2, sheet.max_row + 1):
    production_name = sheet['A' + str(row)].value
    if production_name == 'Garlic':
        sheet['B' + str(row)].value = 1
    elif production_name == 'Celery':
        sheet['B' + str(row)].value = 2
    elif production_name == 'Lemon':
        sheet['B' + str(row)].value = 3
wb.save('C:\Users\RaymondWang\Desktop/produceSales_copy.xlsx')
"""
# More pytorch
Price_Update = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}
for row in range(2, sheet.max_row + 1):
    if sheet.cell(row=row, column=1).value in Price_Update:
        sheet.cell(row=row, column=2).value = Price_Update[sheet.cell(row=row, column=1).value]
wb.save('C:\Users\RaymondWang\Desktop/produceSales_copy.xlsx')
wb.close()
